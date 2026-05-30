"""
MATKA TRACKER v3 - Streamlit Web App
Big upgrade: prediction backtesting, weekday/cycle patterns, market correlations,
better OCR + PDF export, anomaly alerts, edit/delete UI, auto-backup, PWA.
"""
import streamlit as st
import sqlite3, pandas as pd, datetime as dt, calendar, io, os, re, hashlib, shutil
import numpy as np
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

import plotly.express as px
import plotly.graph_objects as go

st.set_page_config(page_title="Matka Tracker v3", page_icon="🎯", layout="wide",
                   initial_sidebar_state="expanded")

DB_PATH = "matka.db"
BACKUP_DIR = "backups"
os.makedirs(BACKUP_DIR, exist_ok=True)

SUM_COLORS = {10:"#4FC3F7",11:"#66BB6A",12:"#FDD835",13:"#FF9800",14:"#EF5350"}
DIGIT_COLORS = ["#4FC3F7","#66BB6A","#FDD835","#FF9800","#EF5350",
                "#7C4DFF","#26C6DA","#FF7043","#9CCC65","#EC407A"]

# Inject PWA support
PWA_HEAD = """
<link rel="manifest" href="./app/static/manifest.json">
<meta name="theme-color" content="#1F4E78">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black">
"""

# ====================================================================
# DB & CORE
# ====================================================================
def get_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_conn(); c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS markets (
        id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL,
        display_order INTEGER NOT NULL, closures TEXT DEFAULT '')""")
    c.execute("""CREATE TABLE IF NOT EXISTS jodis (
        id INTEGER PRIMARY KEY AUTOINCREMENT, market_id INTEGER NOT NULL,
        date TEXT NOT NULL, jodi TEXT NOT NULL,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (market_id) REFERENCES markets(id), UNIQUE(market_id, date))""")
    c.execute("""CREATE TABLE IF NOT EXISTS audit_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT, user TEXT, action TEXT, details TEXT,
        ts TEXT DEFAULT CURRENT_TIMESTAMP)""")
    c.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL, role TEXT DEFAULT 'admin',
        created_at TEXT DEFAULT CURRENT_TIMESTAMP)""")
    c.execute("""CREATE TABLE IF NOT EXISTS predictions_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT, market_id INTEGER NOT NULL,
        prediction_date TEXT NOT NULL, top_jodi TEXT, top_sum INTEGER,
        actual_jodi TEXT, actual_sum INTEGER, top10 TEXT,
        created_at TEXT DEFAULT CURRENT_TIMESTAMP)""")
    if c.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        c.execute("INSERT INTO users (username, password_hash, role) VALUES (?,?,?)",
                  ("admin", hashlib.sha256(b"admin123").hexdigest(), "admin"))
    conn.commit()

def sum_group(j): return (int(j[0]) + int(j[1])) % 5 + 10
def hash_pw(p): return hashlib.sha256(p.encode()).hexdigest()

def log_audit(action, details=""):
    conn = get_conn()
    conn.execute("INSERT INTO audit_log (user, action, details) VALUES (?,?,?)",
                 (st.session_state.get("user","?"), action, details))
    conn.commit()

def auto_backup():
    """Daily auto-backup, keep last 30."""
    today = dt.date.today().isoformat()
    target = f"{BACKUP_DIR}/auto_{today}.db"
    if not os.path.exists(target):
        try:
            shutil.copy(DB_PATH, target)
            # Keep last 30 auto backups
            backups = sorted([f for f in os.listdir(BACKUP_DIR) if f.startswith("auto_")])
            for old in backups[:-30]:
                os.remove(f"{BACKUP_DIR}/{old}")
        except Exception:
            pass

# ====================================================================
# AUTH
# ====================================================================
def login_panel():
    if st.session_state.get("authed"): return True
    st.title("🎯 Matka Tracker v3")
    c1, c2 = st.columns([1,1])
    with c1:
        with st.form("login"):
            st.subheader("Sign in")
            u = st.text_input("Username")
            p = st.text_input("Password", type="password")
            ok = st.form_submit_button("Sign in", use_container_width=True, type="primary")
        if ok:
            row = get_conn().execute("SELECT * FROM users WHERE username=?", (u,)).fetchone()
            if row and row["password_hash"] == hash_pw(p):
                st.session_state["authed"] = True
                st.session_state["user"] = u
                st.session_state["role"] = row["role"]
                log_audit("login")
                st.rerun()
            else:
                st.error("Invalid credentials.")
    with c2:
        st.info("**Default:** admin / admin123\n\nChange in Settings after first login.")
    return False

# ====================================================================
# DATA ACCESS
# ====================================================================
def list_markets():
    return pd.read_sql("SELECT * FROM markets ORDER BY display_order", get_conn())

def add_market(name, closures=""):
    conn = get_conn(); c = conn.cursor()
    mx = c.execute("SELECT COALESCE(MAX(display_order),0) FROM markets").fetchone()[0]
    c.execute("INSERT OR IGNORE INTO markets (name, display_order, closures) VALUES (?,?,?)",
              (name, mx+1, closures))
    conn.commit(); log_audit("add_market", name)

def set_jodi(market_id, date_str, jodi):
    conn = get_conn()
    conn.execute("INSERT OR REPLACE INTO jodis (market_id, date, jodi) VALUES (?,?,?)",
                 (market_id, date_str, jodi.zfill(2)))
    conn.commit()
    log_audit("set_jodi", f"mid={market_id} {date_str}={jodi}")

def delete_jodi(market_id, date_str):
    conn = get_conn()
    conn.execute("DELETE FROM jodis WHERE market_id=? AND date=?", (market_id, date_str))
    conn.commit()
    log_audit("delete_jodi", f"mid={market_id} {date_str}")

def get_jodis(market_id, start=None, end=None):
    q = "SELECT date, jodi FROM jodis WHERE market_id=?"; p = [market_id]
    if start: q += " AND date>=?"; p.append(start)
    if end:   q += " AND date<=?"; p.append(end)
    q += " ORDER BY date"
    return pd.read_sql(q, get_conn(), params=p)

def range_analysis(market_id, start, end):
    df = get_jodis(market_id, start, end)
    sc = {10:0,11:0,12:0,13:0,14:0}; dc = [0]*10
    for _, r in df.iterrows():
        j = r["jodi"]; sc[sum_group(j)] += 1
        dc[int(j[0])] += 1; dc[int(j[1])] += 1
    return {"df":df, "sum_counts":sc, "digit_counts":dc, "total":len(df)}

# ====================================================================
# PREDICTIONS + BACKTEST
# ====================================================================
def build_markov(df):
    """Build Markov from a dataframe of date,jodi (sorted by date)."""
    if len(df) < 5: return None
    jt = defaultdict(Counter); st_ = defaultdict(Counter)
    for i in range(len(df)-1):
        cur, nxt = df.iloc[i]["jodi"], df.iloc[i+1]["jodi"]
        jt[cur][nxt] += 1
        st_[sum_group(cur)][sum_group(nxt)] += 1
    return {"jodi_trans":dict(jt), "sum_trans":dict(st_), "history":df}

def predict_from(df, weekday=None):
    """Return scored predictions given history df."""
    m = build_markov(df)
    if not m: return None
    hist = m["history"]
    last_jodi = hist.iloc[-1]["jodi"]; last_sum = sum_group(last_jodi)
    recent = hist.tail(60)
    sum_freq = Counter(sum_group(j) for j in recent["jodi"])
    digit_freq = Counter()
    for j in recent["jodi"]:
        digit_freq[int(j[0])] += 1; digit_freq[int(j[1])] += 1
    # Day-of-week boost
    weekday_freq = Counter()
    if weekday is not None and "date" in hist.columns:
        for _, r in hist.tail(180).iterrows():
            d = pd.to_datetime(r["date"])
            if d.weekday() == weekday:
                weekday_freq[sum_group(r["jodi"])] += 1
    mn = m["jodi_trans"].get(last_jodi, Counter())
    smn = m["sum_trans"].get(last_sum, Counter())
    scores = {}
    for n in range(100):
        j = f"{n:02d}"; sg = sum_group(j); s = 0.0
        s += 3.0 * (mn.get(j,0) / (sum(mn.values()) or 1))
        s += 1.5 * (smn.get(sg,0) / (sum(smn.values()) or 1))
        s += 1.0 * (sum_freq.get(sg,0) / (sum(sum_freq.values()) or 1))
        s += 0.5 * ((digit_freq.get(int(j[0]),0)+digit_freq.get(int(j[1]),0)) /
                    (sum(digit_freq.values()) or 1))
        if weekday_freq:
            s += 1.0 * (weekday_freq.get(sg,0) / (sum(weekday_freq.values()) or 1))
        scores[j] = s
    top = sorted(scores.items(), key=lambda x: -x[1])[:10]
    return {"top":top, "last_jodi":last_jodi, "last_sum":last_sum,
            "markov_size":len(m["jodi_trans"]),
            "sum_freq":dict(sum_freq), "digit_freq":dict(digit_freq.most_common(10))}

def predict_for_market(market_id, weekday=None):
    df = pd.read_sql("SELECT date, jodi FROM jodis WHERE market_id=? ORDER BY date",
                     get_conn(), params=[market_id])
    return predict_from(df, weekday)

def backtest_market(market_id, lookback=30):
    """Walk-forward backtest: predict each of the last `lookback` days and check hit rate."""
    df = pd.read_sql("SELECT date, jodi FROM jodis WHERE market_id=? ORDER BY date",
                     get_conn(), params=[market_id])
    if len(df) < lookback + 20:
        return None
    results = []
    for i in range(len(df) - lookback, len(df)):
        train = df.iloc[:i].copy().reset_index(drop=True)
        actual = df.iloc[i]
        weekday = pd.to_datetime(actual["date"]).weekday()
        pred = predict_from(train, weekday)
        if not pred: continue
        top1 = pred["top"][0][0]
        top3 = [j for j, _ in pred["top"][:3]]
        top10 = [j for j, _ in pred["top"][:10]]
        actual_jodi = actual["jodi"]
        actual_sum = sum_group(actual_jodi)
        predicted_sum = sum_group(top1)
        results.append({
            "date": actual["date"],
            "actual_jodi": actual_jodi,
            "actual_sum": actual_sum,
            "top1_jodi": top1,
            "top1_sum": predicted_sum,
            "exact_hit_top1": actual_jodi == top1,
            "exact_hit_top3": actual_jodi in top3,
            "exact_hit_top10": actual_jodi in top10,
            "sum_hit": actual_sum == predicted_sum,
        })
    return pd.DataFrame(results)

# ====================================================================
# PATTERN DISCOVERY
# ====================================================================
def weekday_pattern(market_id):
    df = pd.read_sql("SELECT date, jodi FROM jodis WHERE market_id=? ORDER BY date",
                     get_conn(), params=[market_id])
    if df.empty: return None
    df["date"] = pd.to_datetime(df["date"])
    df["weekday"] = df["date"].dt.day_name()
    df["sum"] = df["jodi"].apply(sum_group)
    pivot = df.groupby(["weekday","sum"]).size().unstack(fill_value=0)
    for sg in [10,11,12,13,14]:
        if sg not in pivot.columns: pivot[sg] = 0
    pivot = pivot[[10,11,12,13,14]]
    pivot = pivot.reindex(["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"])
    return pivot

def cycle_detection(market_id, max_lag=14):
    """Find autocorrelation peaks in sum group sequence — indicates cyclical patterns."""
    df = pd.read_sql("SELECT date, jodi FROM jodis WHERE market_id=? ORDER BY date",
                     get_conn(), params=[market_id])
    if len(df) < 30: return None
    sums = df["jodi"].apply(sum_group).values
    # Convert to one-hot per sum group, sum cross-correlation
    correlations = []
    for lag in range(1, max_lag+1):
        if lag >= len(sums): break
        a = sums[:-lag]; b = sums[lag:]
        match_rate = float((a == b).mean())
        correlations.append({"Lag (days)": lag, "Match rate": round(match_rate, 3)})
    return pd.DataFrame(correlations)

def market_correlation(market_ids, target_sum=14):
    """For each pair of markets, what % of days they both hit `target_sum`."""
    conn = get_conn()
    results = []
    market_data = {}
    for mid in market_ids:
        df = pd.read_sql("SELECT date, jodi FROM jodis WHERE market_id=?", conn, params=[mid])
        if df.empty: continue
        df["sg"] = df["jodi"].apply(sum_group)
        market_data[mid] = df[df["sg"] == target_sum]["date"].tolist()
    name_lookup = {row["id"]: row["name"] for _, row in list_markets().iterrows()}
    for i, mi in enumerate(market_ids):
        for mj in market_ids[i+1:]:
            si = set(market_data.get(mi, [])); sj = set(market_data.get(mj, []))
            if not si or not sj: continue
            inter = len(si & sj); union = len(si | sj)
            j_idx = inter / union if union else 0
            results.append({
                "Market A": name_lookup[mi], "Market B": name_lookup[mj],
                f"Both Sum {target_sum} days": inter,
                "Either day": union, "Jaccard similarity": round(j_idx, 3)
            })
    return pd.DataFrame(results).sort_values("Jaccard similarity", ascending=False) if results else None

def detect_anomalies(market_id):
    """Flag unusual patterns: never-before jodis, long streaks, etc."""
    df = pd.read_sql("SELECT date, jodi FROM jodis WHERE market_id=? ORDER BY date",
                     get_conn(), params=[market_id])
    if df.empty: return []
    anomalies = []
    df["sg"] = df["jodi"].apply(sum_group)
    # Long streaks
    df["run"] = (df["sg"] != df["sg"].shift()).cumsum()
    streaks = df.groupby(["run","sg"]).size()
    max_streak = streaks.max() if len(streaks) else 0
    if max_streak >= 5:
        for (run, sg), cnt in streaks.items():
            if cnt >= 5:
                start_idx = df[df["run"] == run].iloc[0].name
                end_idx = df[df["run"] == run].iloc[-1].name
                anomalies.append({
                    "Type": "Long streak",
                    "Description": f"Sum {sg} for {cnt} consecutive days",
                    "Date": f"{df.loc[start_idx,'date']} → {df.loc[end_idx,'date']}"
                })
    # Most common jodi (interesting if very dominant)
    jc = df["jodi"].value_counts()
    if len(df) > 30 and jc.iloc[0] > len(df) * 0.05:
        anomalies.append({
            "Type": "Dominant jodi",
            "Description": f"Jodi {jc.index[0]} appeared {jc.iloc[0]} times ({jc.iloc[0]/len(df)*100:.1f}%)",
            "Date": "All time"
        })
    return anomalies

# ====================================================================
# OCR (multi-strategy)
# ====================================================================
def try_ocr(uploaded_file):
    """Try multiple OCR strategies."""
    try:
        from PIL import Image
        img = Image.open(uploaded_file)
    except Exception as e:
        return {"ok": False, "error": f"Could not open image: {e}"}
    # Strategy 1: pytesseract
    try:
        import pytesseract
        text = pytesseract.image_to_string(img)
        nums = re.findall(r'\b\d{2}\b', text)
        return {"ok": True, "img": img, "numbers": nums, "raw_text": text, "method": "pytesseract"}
    except ImportError:
        pass
    except Exception:
        pass
    # Strategy 2: easyocr
    try:
        import easyocr
        reader = easyocr.Reader(['en'])
        import numpy as np
        text = " ".join(reader.readtext(np.array(img), detail=0))
        nums = re.findall(r'\b\d{2}\b', text)
        return {"ok": True, "img": img, "numbers": nums, "raw_text": text, "method": "easyocr"}
    except ImportError:
        pass
    except Exception:
        pass
    return {"ok": False, "img": img,
            "error": "No OCR engine available. Install one:\n"
                     "**Option 1 (recommended):** `pip install pytesseract` + install tesseract: https://github.com/UB-Mannheim/tesseract/wiki\n"
                     "**Option 2:** `pip install easyocr` (slow first run, downloads model)"}

# ====================================================================
# EXCEL & PDF
# ====================================================================
def build_range_excel(market_ids, sd, ed, title="Custom Range"):
    if not market_ids: return None
    conn = get_conn()
    mks = pd.read_sql(
        f"SELECT * FROM markets WHERE id IN ({','.join(['?']*len(market_ids))}) ORDER BY display_order",
        conn, params=market_ids)
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    thin = Side(border_style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hf = PatternFill("solid", fgColor="305496")
    hfo = Font(bold=True, color="FFFFFF", size=10)
    tf = Font(bold=True, color="FFFFFF", size=13)
    tfi = PatternFill("solid", fgColor="1F4E78")
    cen = Alignment(horizontal="center", vertical="center")
    la = Alignment(horizontal="left", vertical="center")
    SF = {10:PatternFill("solid",fgColor="D9EEFB"),11:PatternFill("solid",fgColor="DDEFD8"),
          12:PatternFill("solid",fgColor="FFF5CC"),13:PatternFill("solid",fgColor="FCE4CC"),
          14:PatternFill("solid",fgColor="FAD4D4")}
    SFC = {10:"1F6FA5",11:"3E8B45",12:"9C7700",13:"B05D00",14:"B23030"}
    DH = ["4FC3F7","66BB6A","FDD835","FF9800","EF5350","7C4DFF","26C6DA","FF7043","9CCC65","EC407A"]
    def styh(c):
        c.fill=hf; c.font=hfo; c.alignment=cen; c.border=border
    def stysum(c, sv):
        c.alignment=cen; c.border=border
        if sv in SF: c.fill=SF[sv]; c.font=Font(bold=True, color=SFC[sv])
    ws = wb.create_sheet("Sum 10-14")
    ws.cell(row=1, column=1, value=f"{title} ({sd} to {ed})").font = tf
    ws.cell(row=1, column=1).fill = tfi; ws.cell(row=1, column=1).alignment = cen
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    for j, h in enumerate(["Market","Days","Sum 10","Sum 11","Sum 12","Sum 13","Sum 14","Total","Hottest","Coldest"]):
        styh(ws.cell(row=3, column=1+j, value=h))
    ov = {10:0,11:0,12:0,13:0,14:0}
    for i, m in mks.iterrows():
        r = 4+i
        c = ws.cell(row=r, column=1, value=m["name"]); c.border=border; c.alignment=la
        c.font = Font(bold=True)
        a = range_analysis(m["id"], sd, ed)
        ws.cell(row=r, column=2, value=a["total"]).alignment = cen
        ws.cell(row=r, column=2).border = border
        for k, sg in enumerate([10,11,12,13,14]):
            stysum(ws.cell(row=r, column=3+k, value=a["sum_counts"][sg]), sg)
            ov[sg] += a["sum_counts"][sg]
        tot = sum(a["sum_counts"].values())
        tc = ws.cell(row=r, column=8, value=tot); tc.alignment=cen; tc.border=border; tc.font=Font(bold=True)
        if tot:
            mx = max(a["sum_counts"].values()); mn = min(a["sum_counts"].values())
            hot = ", ".join(str(s) for s in [10,11,12,13,14] if a["sum_counts"][s]==mx)
            cold = ", ".join(str(s) for s in [10,11,12,13,14] if a["sum_counts"][s]==mn)
        else: hot = cold = "-"
        ws.cell(row=r, column=9, value=hot).alignment = cen; ws.cell(row=r, column=9).border = border
        ws.cell(row=r, column=10, value=cold).alignment = cen; ws.cell(row=r, column=10).border = border
    r = 4 + len(mks)
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor="D9D9D9")
    ws.cell(row=r, column=1).border = border
    for k, sg in enumerate([10,11,12,13,14]):
        c = ws.cell(row=r, column=3+k, value=ov[sg])
        stysum(c, sg); c.font = Font(bold=True, color=SFC[sg], size=11)
    ws.cell(row=r, column=8, value=sum(ov.values())).font = Font(bold=True)
    ws.cell(row=r, column=8).alignment = cen; ws.cell(row=r, column=8).border = border
    ws.column_dimensions['A'].width = 22; ws.column_dimensions['B'].width = 7
    for col in range(3, 11): ws.column_dimensions[get_column_letter(col)].width = 11
    chart = BarChart(); chart.type="bar"; chart.grouping="stacked"; chart.overlap=100
    chart.title = "Sum 10-14 per Market"
    chart.height=22; chart.width=24
    chart.add_data(Reference(ws, min_col=3, min_row=3, max_row=3+len(mks), max_col=7), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=4, max_row=3+len(mks)))
    ws.add_chart(chart, "L3")
    ws2 = wb.create_sheet("Digit 0-9")
    for j, h in enumerate(["Market","Days"] + [f"D{d}" for d in range(10)] + ["Total"]):
        styh(ws2.cell(row=3, column=1+j, value=h))
    for i, m in mks.iterrows():
        r = 4+i
        ws2.cell(row=r, column=1, value=m["name"]).font = Font(bold=True)
        ws2.cell(row=r, column=1).border = border
        a = range_analysis(m["id"], sd, ed)
        ws2.cell(row=r, column=2, value=a["total"]).alignment = cen
        for k in range(10):
            c = ws2.cell(row=r, column=3+k, value=a["digit_counts"][k])
            c.alignment=cen; c.border=border
            c.fill = PatternFill("solid", fgColor=DH[k]); c.font = Font(bold=True)
        ws2.cell(row=r, column=13, value=sum(a["digit_counts"])).font = Font(bold=True)
        ws2.cell(row=r, column=13).alignment = cen; ws2.cell(row=r, column=13).border = border
    ws2.column_dimensions['A'].width = 22
    for col in range(2, 14): ws2.column_dimensions[get_column_letter(col)].width = 7
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

def build_pdf_report(market_ids, sd, ed, title="Report"):
    """PDF version of the range report."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
    except ImportError:
        return None
    if not market_ids: return None
    conn = get_conn()
    mks = pd.read_sql(
        f"SELECT * FROM markets WHERE id IN ({','.join(['?']*len(market_ids))}) ORDER BY display_order",
        conn, params=market_ids)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph(f"<b>{title}</b>", styles['Title']))
    elements.append(Paragraph(f"{sd} to {ed}", styles['Normal']))
    elements.append(Spacer(1, 12))
    # Sum 10-14 table
    data = [["Market", "Days", "Sum 10", "Sum 11", "Sum 12", "Sum 13", "Sum 14", "Total"]]
    sum_colors = {10:colors.HexColor("#D9EEFB"), 11:colors.HexColor("#DDEFD8"),
                  12:colors.HexColor("#FFF5CC"), 13:colors.HexColor("#FCE4CC"),
                  14:colors.HexColor("#FAD4D4")}
    overall = {10:0,11:0,12:0,13:0,14:0}
    for _, m in mks.iterrows():
        a = range_analysis(m["id"], sd, ed)
        row = [m["name"], a["total"]] + [a["sum_counts"][s] for s in [10,11,12,13,14]] + [sum(a["sum_counts"].values())]
        data.append(row)
        for s in [10,11,12,13,14]: overall[s] += a["sum_counts"][s]
    data.append(["TOTAL", "", overall[10], overall[11], overall[12], overall[13], overall[14], sum(overall.values())])
    t = Table(data, colWidths=[5*cm, 1.5*cm] + [2*cm]*5 + [1.8*cm])
    style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#305496")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (1,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor("#D9D9D9")),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
    ])
    for col_idx, sg in enumerate([10,11,12,13,14]):
        style.add('BACKGROUND', (2+col_idx,1), (2+col_idx,-2), sum_colors[sg])
    t.setStyle(style)
    elements.append(t)
    doc.build(elements)
    buf.seek(0)
    return buf

def build_single_market_excel(market_id):
    df = get_jodis(market_id)
    if df.empty: return None
    df["sum"] = df["jodi"].apply(sum_group)
    df["date"] = pd.to_datetime(df["date"])
    df["month"] = df["date"].dt.strftime("%Y-%m")
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    SF = {10:PatternFill("solid",fgColor="D9EEFB"),11:PatternFill("solid",fgColor="DDEFD8"),
          12:PatternFill("solid",fgColor="FFF5CC"),13:PatternFill("solid",fgColor="FCE4CC"),
          14:PatternFill("solid",fgColor="FAD4D4")}
    ws = wb.create_sheet("History")
    ws.append(["Date","Day","Jodi","Sum"])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
    for _, r in df.iterrows():
        ws.append([r["date"].strftime("%Y-%m-%d"), r["date"].strftime("%a"), r["jodi"], r["sum"]])
        ws.cell(row=ws.max_row, column=4).fill = SF[r["sum"]]
    ws.column_dimensions['A'].width = 14
    for c in 'BCD': ws.column_dimensions[c].width = 10
    ws2 = wb.create_sheet("Monthly Summary")
    monthly = df.groupby(["month","sum"]).size().unstack(fill_value=0)
    for sg in [10,11,12,13,14]:
        if sg not in monthly.columns: monthly[sg] = 0
    monthly = monthly[[10,11,12,13,14]]
    monthly["Total"] = monthly.sum(axis=1)
    ws2.append(["Month"] + [f"Sum {s}" for s in [10,11,12,13,14]] + ["Total"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="305496")
    for mo, row in monthly.iterrows():
        ws2.append([mo] + [int(row[s]) for s in [10,11,12,13,14]] + [int(row["Total"])])
    ws2.column_dimensions['A'].width = 12
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ====================================================================
# PAGES
# ====================================================================
def page_dashboard():
    st.title("📊 Live Dashboard")
    md = list_markets()
    if md.empty: st.warning("No markets."); return
    c1, c2, c3, _ = st.columns([1,1,1,2])
    today = dt.date.today()
    with c1:
        preset = st.selectbox("Range", ["This Month","Last 7 Days","Last 30 Days","Custom"])
    if preset == "This Month": start, end = today.replace(day=1), today
    elif preset == "Last 7 Days": start, end = today - dt.timedelta(days=6), today
    elif preset == "Last 30 Days": start, end = today - dt.timedelta(days=29), today
    else:
        with c2: start = st.date_input("From", value=today - dt.timedelta(days=30))
        with c3: end = st.date_input("To", value=today)
    ss, es = start.isoformat(), end.isoformat()
    ov = {10:0,11:0,12:0,13:0,14:0}; tot = 0; dc = [0]*10
    for _, m in md.iterrows():
        a = range_analysis(m["id"], ss, es)
        for s in [10,11,12,13,14]: ov[s] += a["sum_counts"][s]
        for d in range(10): dc[d] += a["digit_counts"][d]
        tot += a["total"]
    st.subheader(f"{ss} → {es} · {len(md)} markets")
    cols = st.columns(7)
    cols[0].metric("Total Jodis", tot)
    cols[1].metric("Markets", len(md))
    for i, s in enumerate([10,11,12,13,14]):
        cols[2+i].metric(f"Sum {s}", ov[s])
    if tot == 0: st.info("No data."); return
    c1, c2 = st.columns(2)
    with c1:
        sdf = pd.DataFrame({"Sum":[f"Sum {s}" for s in [10,11,12,13,14]],
                           "Count":[ov[s] for s in [10,11,12,13,14]]})
        fig = px.bar(sdf, x="Sum", y="Count", color="Sum",
                    color_discrete_map={f"Sum {s}":SUM_COLORS[s] for s in [10,11,12,13,14]},
                    title="Sum 10-14")
        fig.update_layout(height=340, showlegend=False)
        st.plotly_chart(fig, use_container_width=True)
    with c2:
        ddf = pd.DataFrame({"Digit":[str(d) for d in range(10)], "Count":dc})
        fig2 = px.bar(ddf, x="Digit", y="Count", color="Digit",
                     color_discrete_map={str(d):DIGIT_COLORS[d] for d in range(10)},
                     title="Digit 0-9")
        fig2.update_layout(height=340, showlegend=False)
        st.plotly_chart(fig2, use_container_width=True)
    # Digit 0-9 total count table
    st.subheader("🔢 Digit 0-9 Total Count")
    grand = sum(dc) or 1
    hot_d = dc.index(max(dc))
    cold_d = dc.index(min(dc))
    digit_cols = st.columns(11)
    digit_cols[0].markdown("**Digit**")
    for d in range(10):
        digit_cols[d+1].markdown(f"<div style='text-align:center;background:{DIGIT_COLORS[d]};color:white;padding:8px;border-radius:6px;font-weight:bold;font-size:18px'>{d}</div>", unsafe_allow_html=True)
    digit_cols2 = st.columns(11)
    digit_cols2[0].markdown("**Count**")
    for d in range(10):
        digit_cols2[d+1].markdown(f"<div style='text-align:center;padding:6px;font-weight:bold;font-size:16px'>{dc[d]}</div>", unsafe_allow_html=True)
    digit_cols3 = st.columns(11)
    digit_cols3[0].markdown("**%**")
    for d in range(10):
        pct = dc[d] / grand * 100
        digit_cols3[d+1].markdown(f"<div style='text-align:center;padding:4px;color:#888'>{pct:.1f}%</div>", unsafe_allow_html=True)
    cA, cB, cC = st.columns(3)
    cA.metric("🔥 Hottest Digit", f"{hot_d}", f"{dc[hot_d]} plays")
    cB.metric("❄️ Coldest Digit", f"{cold_d}", f"{dc[cold_d]} plays")
    cC.metric("📊 Total Digit Plays", grand, f"avg {grand/10:.0f}/digit")

    st.subheader("🔥 Current Streaks (date range)")
    streaks = []
    for _, m in md.iterrows():
        df = get_jodis(m["id"], ss, es)
        if df.empty: continue
        df["sg"] = df["jodi"].apply(sum_group)
        last_sg = df.iloc[-1]["sg"]; streak = 1
        for i in range(len(df)-2, -1, -1):
            if df.iloc[i]["sg"] == last_sg: streak += 1
            else: break
        streaks.append({"Market":m["name"], "Last Sum":int(last_sg), "Streak":streak,
                       "Last Jodi":df.iloc[-1]["jodi"], "Last Date":df.iloc[-1]["date"]})
    if streaks:
        st.dataframe(pd.DataFrame(streaks).sort_values("Streak", ascending=False),
                    use_container_width=True, hide_index=True)
    st.subheader("🚨 Anomalies (across all markets, all-time)")
    anom_rows = []
    for _, m in md.iterrows():
        for a in detect_anomalies(m["id"]):
            anom_rows.append({"Market": m["name"], **a})
    if anom_rows:
        st.dataframe(pd.DataFrame(anom_rows), use_container_width=True, hide_index=True)
    else:
        st.info("No notable anomalies.")
    st.subheader("📝 Recent Activity")
    rec = pd.read_sql("SELECT ts, user, action, details FROM audit_log ORDER BY id DESC LIMIT 20",
                     get_conn())
    st.dataframe(rec, use_container_width=True, hide_index=True, height=240)

def page_data_entry():
    st.title("📝 Data Entry")
    md = list_markets()
    if md.empty: st.warning("Add markets first."); return
    tabs = st.tabs(["⚡ Quick Today","✏️ Single","📋 Bulk","📁 CSV","📷 Photo OCR","✂️ Edit/Delete"])
    with tabs[0]:
        st.subheader("Enter today's jodis")
        today = dt.date.today(); entries = {}
        cc = 3; n = len(md); rpc = (n + cc - 1) // cc
        cols = st.columns(cc)
        for i, m in md.iterrows():
            col = cols[i // rpc]
            with col:
                v = st.text_input(m["name"], max_chars=2, key=f"qt_{m['id']}", placeholder="00")
                if v: entries[m["id"]] = v
        if st.button(f"💾 Save all for {today}", type="primary", use_container_width=True):
            ok = err = 0
            for mid, v in entries.items():
                if len(v)==2 and v.isdigit(): set_jodi(mid, today.isoformat(), v); ok+=1
                else: err+=1
            if ok: st.success(f"Saved {ok}")
            if err: st.warning(f"{err} skipped")
            if ok:
                for k in list(st.session_state.keys()):
                    if k.startswith("qt_"): del st.session_state[k]
                st.rerun()
    with tabs[1]:
        st.subheader("Single Entry")
        c1, c2, c3 = st.columns(3)
        with c1: market = st.selectbox("Market", md["name"].tolist())
        with c2: date = st.date_input("Date", value=dt.date.today())
        with c3: jodi_in = st.text_input("Jodi (0-99)", max_chars=2, placeholder="45 or 5")
        if st.button("Save", type="primary", key="se"):
            # Accept 1 or 2 digit input, auto-pad
            jodi_clean = (jodi_in or "").strip()
            if jodi_clean and jodi_clean.isdigit() and 0 <= int(jodi_clean) <= 99:
                jodi_padded = jodi_clean.zfill(2)
                mid = int(md[md["name"]==market]["id"].iloc[0])
                set_jodi(mid, date.isoformat(), jodi_padded)
                st.success(f"✓ Saved {market} {date}: jodi {jodi_padded} → Sum {sum_group(jodi_padded)}")
            else:
                st.error("Enter a number 0-99 (e.g. 45, or 5 for 05).")
    with tabs[2]:
        st.subheader("Bulk paste")
        st.code("Sita Day, 2026-06-01, 45")
        bulk = st.text_area("Paste lines", height=200)
        if st.button("Import", type="primary", key="bi"):
            ok = err = 0; errs = []
            for ln in [l.strip() for l in bulk.split("\n") if l.strip()]:
                parts = [p.strip() for p in ln.split(",")]
                if len(parts) != 3: err+=1; errs.append(f"Bad: {ln}"); continue
                m, d, j = parts
                if m not in md["name"].tolist(): err+=1; errs.append(f"Unknown: {m}"); continue
                try: dt.date.fromisoformat(d)
                except: err+=1; errs.append(f"Bad date: {d}"); continue
                if not (len(j)==2 and j.isdigit()): err+=1; errs.append(f"Bad jodi: {j}"); continue
                mid = int(md[md["name"]==m]["id"].iloc[0])
                set_jodi(mid, d, j); ok+=1
            st.success(f"Imported {ok}")
            if err:
                st.error(f"{err} errors")
                for e in errs[:20]: st.text(e)
    with tabs[3]:
        st.subheader("CSV Upload")
        up = st.file_uploader("CSV/TSV", type=["csv","tsv"])
        if up:
            sep = "\t" if up.name.endswith(".tsv") else ","
            try:
                df = pd.read_csv(up, sep=sep)
                df.columns = [c.strip().lower() for c in df.columns]
                st.dataframe(df.head(10))
                if st.button("Import CSV", type="primary"):
                    ok = err = 0
                    for _, r in df.iterrows():
                        m = str(r.get("market","")).strip()
                        d = str(r.get("date","")).strip()
                        j = str(r.get("jodi","")).strip().zfill(2)
                        if m in md["name"].tolist() and len(j)==2 and j.isdigit():
                            try:
                                dt.date.fromisoformat(d)
                                mid = int(md[md["name"]==m]["id"].iloc[0])
                                set_jodi(mid, d, j); ok+=1
                            except: err+=1
                        else: err+=1
                    st.success(f"Imported {ok}, skipped {err}")
            except Exception as e:
                st.error(f"Error: {e}")
    with tabs[4]:
        st.subheader("Photo OCR")
        st.caption("Multi-engine: tries pytesseract first, then easyocr if installed.")
        img_up = st.file_uploader("Chart image", type=["png","jpg","jpeg"])
        if img_up:
            r = try_ocr(img_up)
            if not r["ok"]:
                st.error(r["error"])
                if "img" in r: st.image(r["img"], use_container_width=True)
            else:
                st.success(f"Extracted via **{r['method']}**")
                st.image(r["img"], use_container_width=True)
                st.write(f"**Detected 2-digit numbers:** {r['numbers']}")
                st.text_area("Raw OCR", r["raw_text"], height=150)
                st.warning("Verify each before importing.")
    with tabs[5]:
        st.subheader("Edit / Delete entries")
        market = st.selectbox("Market", md["name"].tolist(), key="ed_m")
        mid = int(md[md["name"]==market]["id"].iloc[0])
        c1, c2 = st.columns(2)
        today = dt.date.today()
        with c1: ed_start = st.date_input("From", value=today - dt.timedelta(days=30), key="ed_s")
        with c2: ed_end = st.date_input("To", value=today, key="ed_e")
        df = get_jodis(mid, ed_start.isoformat(), ed_end.isoformat())
        if df.empty: st.info("No data in range.")
        else:
            df["sum"] = df["jodi"].apply(sum_group)
            st.dataframe(df, use_container_width=True, hide_index=True)
            del_date = st.selectbox("Date to delete", df["date"].tolist())
            if st.button("🗑️ Delete entry", type="secondary"):
                delete_jodi(mid, del_date)
                st.success(f"Deleted {market} {del_date}")
                st.rerun()

def page_predictions():
    st.title("🔮 Predictions")
    st.caption("⚠️ Statistical only. Not gambling advice.")
    md = list_markets()
    market = st.selectbox("Market", md["name"].tolist())
    mid = int(md[md["name"]==market]["id"].iloc[0])
    next_date = dt.date.today() + dt.timedelta(days=1)
    weekday = st.date_input("Predict for date", value=next_date).weekday()
    r = predict_for_market(mid, weekday=weekday)
    if not r:
        st.warning("Need at least 5 entries."); return
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Last Jodi", r["last_jodi"])
    c2.metric("Last Sum", f"Sum {r['last_sum']}")
    c3.metric("Unique Jodis Seen", r["markov_size"])
    c4.metric("Recent Days", sum(r["sum_freq"].values()))
    st.subheader("🎯 Top 10 Suggested Jodis")
    st.caption("Score = Markov(3x) + Sum-Markov(1.5x) + Recent Sum freq(1x) + Digit freq(0.5x) + Weekday(1x)")
    td = pd.DataFrame(r["top"], columns=["Jodi","Score"])
    td["Sum"] = td["Jodi"].apply(sum_group)
    td["Score"] = td["Score"].apply(lambda x: f"{x:.3f}")
    def cr(row):
        return [f"background-color: {SUM_COLORS[int(row['Sum'])]}; color: white; font-weight: bold" for _ in row]
    st.dataframe(td.style.apply(cr, axis=1), use_container_width=True, hide_index=True)

def page_backtest():
    st.title("🧪 Prediction Backtest")
    st.caption("Walk-forward test: pretend today is each of the last N days, predict, check accuracy.")
    md = list_markets()
    market = st.selectbox("Market", md["name"].tolist())
    mid = int(md[md["name"]==market]["id"].iloc[0])
    lb = st.slider("Days to backtest", 10, 60, 30)
    if st.button("Run Backtest", type="primary"):
        with st.spinner(f"Backtesting last {lb} days..."):
            res = backtest_market(mid, lb)
        if res is None or len(res) == 0:
            st.error("Not enough history.")
            return
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Days tested", len(res))
        c2.metric("Top-1 exact hit", f"{res['exact_hit_top1'].mean()*100:.1f}%")
        c3.metric("Top-10 exact hit", f"{res['exact_hit_top10'].mean()*100:.1f}%")
        c4.metric("Sum group hit", f"{res['sum_hit'].mean()*100:.1f}%")
        st.caption(f"Baseline if random: Top-1=1%, Top-10=10%, Sum=20%")
        st.subheader("Per-day results")
        def color_hits(row):
            if row["exact_hit_top1"]: c = "#66BB6A"
            elif row["exact_hit_top3"]: c = "#FDD835"
            elif row["exact_hit_top10"]: c = "#FF9800"
            else: c = "transparent"
            return [f"background-color: {c}" for _ in row]
        display = res[["date","actual_jodi","actual_sum","top1_jodi","top1_sum",
                      "exact_hit_top1","exact_hit_top3","exact_hit_top10","sum_hit"]]
        st.dataframe(display.style.apply(color_hits, axis=1), use_container_width=True, hide_index=True)

def page_patterns():
    st.title("🧩 Pattern Discovery")
    md = list_markets()
    market = st.selectbox("Market", md["name"].tolist())
    mid = int(md[md["name"]==market]["id"].iloc[0])
    tabs = st.tabs(["📅 Weekday Patterns","♻️ Cycles","🔗 Market Correlations"])
    with tabs[0]:
        st.subheader("Sum 10-14 by Day-of-Week")
        pat = weekday_pattern(mid)
        if pat is None:
            st.warning("No data.")
        else:
            fig = px.imshow(pat.values,
                           labels=dict(x="Sum Group", y="Weekday", color="Count"),
                           x=[f"Sum {s}" for s in [10,11,12,13,14]],
                           y=pat.index.tolist(),
                           color_continuous_scale="YlOrRd", aspect="auto",
                           title="Weekday × Sum heatmap")
            fig.update_layout(height=400)
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(pat, use_container_width=True)
    with tabs[1]:
        st.subheader("Cycle Detection (autocorrelation)")
        st.caption("Higher 'Match rate' = sum group tends to repeat after that many days.")
        cyc = cycle_detection(mid)
        if cyc is None:
            st.warning("Need 30+ days of data.")
        else:
            fig = px.line(cyc, x="Lag (days)", y="Match rate", markers=True,
                         title="Sum-group autocorrelation")
            baseline = 0.20  # random baseline
            fig.add_hline(y=baseline, line_dash="dash", line_color="gray",
                         annotation_text="Random baseline (20%)")
            st.plotly_chart(fig, use_container_width=True)
            st.dataframe(cyc.sort_values("Match rate", ascending=False),
                        use_container_width=True, hide_index=True)
    with tabs[2]:
        st.subheader("Market Correlations")
        target_sum = st.selectbox("Target sum", [10,11,12,13,14], index=4)
        sel = st.multiselect("Markets to compare", md["name"].tolist(),
                            default=md["name"].tolist()[:6])
        if len(sel) >= 2:
            sel_ids = md[md["name"].isin(sel)]["id"].tolist()
            corr = market_correlation(sel_ids, target_sum)
            if corr is None or corr.empty:
                st.info("No overlapping dates.")
            else:
                st.dataframe(corr, use_container_width=True, hide_index=True)
                st.caption("Jaccard = (both same sum) / (either day). Higher = markets tend to hit this sum on same days.")

def page_reports():
    st.title("📑 Reports & Export")
    md = list_markets()
    tabs = st.tabs(["📅 Date Range Excel","📄 PDF Report","🎯 Single Market","📊 Comparison"])
    with tabs[0]:
        st.subheader("Custom Date Range Excel")
        today = dt.date.today()
        c1, c2, c3 = st.columns(3)
        with c1: start = st.date_input("From", value=today.replace(day=1))
        with c2: end = st.date_input("To", value=today)
        with c3: sel = st.multiselect("Markets", md["name"].tolist(), default=md["name"].tolist())
        if st.button("Generate Excel", type="primary"):
            ids = md[md["name"].isin(sel)]["id"].tolist()
            buf = build_range_excel(ids, start.isoformat(), end.isoformat())
            if buf:
                st.download_button("⬇️ Download Excel", data=buf,
                                  file_name=f"Matka_{start}_to_{end}.xlsx",
                                  mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with tabs[1]:
        st.subheader("PDF Report")
        today = dt.date.today()
        c1, c2, c3 = st.columns(3)
        with c1: pstart = st.date_input("From", value=today.replace(day=1), key="ps")
        with c2: pend = st.date_input("To", value=today, key="pe")
        with c3: psel = st.multiselect("Markets", md["name"].tolist(),
                                      default=md["name"].tolist()[:10], key="psel")
        if st.button("Generate PDF", type="primary"):
            ids = md[md["name"].isin(psel)]["id"].tolist()
            buf = build_pdf_report(ids, pstart.isoformat(), pend.isoformat(), title="Matka Report")
            if buf is None:
                st.error("PDF generation requires `pip install reportlab`. After install, restart app.")
            else:
                st.download_button("⬇️ Download PDF", data=buf,
                                  file_name=f"Matka_{pstart}_to_{pend}.pdf",
                                  mime="application/pdf")
    with tabs[2]:
        st.subheader("Single Market Full History Excel")
        m = st.selectbox("Market", md["name"].tolist(), key="sm")
        mid = int(md[md["name"]==m]["id"].iloc[0])
        if st.button("Generate", type="primary", key="smg"):
            buf = build_single_market_excel(mid)
            if buf:
                st.download_button("⬇️ Download", data=buf,
                                  file_name=f"{m.replace(' ','_')}_History.xlsx",
                                  mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else: st.error("No data.")
    with tabs[3]:
        st.subheader("Multi-Market Comparison")
        sel = st.multiselect("Markets", md["name"].tolist(), default=md["name"].tolist()[:3], key="cs")
        today = dt.date.today()
        c1, c2 = st.columns(2)
        with c1: start = st.date_input("From", value=today - dt.timedelta(days=60), key="csd")
        with c2: end = st.date_input("To", value=today, key="ced")
        if sel:
            rows = []
            for n in sel:
                m = md[md["name"]==n].iloc[0]
                df = get_jodis(int(m["id"]), start.isoformat(), end.isoformat())
                if df.empty: continue
                df["date"] = pd.to_datetime(df["date"])
                df["month"] = df["date"].dt.strftime("%Y-%m")
                df["sg"] = df["jodi"].apply(sum_group)
                for mo, grp in df.groupby("month"):
                    sc = grp["sg"].value_counts().to_dict()
                    rows.append({"Market":n, "Month":mo,
                                **{f"Sum {s}":int(sc.get(s,0)) for s in [10,11,12,13,14]},
                                "Total":len(grp)})
            if rows:
                cdf = pd.DataFrame(rows)
                st.dataframe(cdf, use_container_width=True, hide_index=True)
                for s in [10,11,12,13,14]:
                    fig = px.line(cdf, x="Month", y=f"Sum {s}", color="Market",
                                 markers=True, title=f"Sum {s}")
                    st.plotly_chart(fig, use_container_width=True)

def page_market_history():
    st.title("📈 Market History")
    md = list_markets()
    market = st.selectbox("Market", md["name"].tolist())
    mid = int(md[md["name"]==market]["id"].iloc[0])
    df = get_jodis(mid)
    if df.empty: st.info("No data."); return
    df["date"] = pd.to_datetime(df["date"])
    df["sum"] = df["jodi"].apply(sum_group)
    df["month"] = df["date"].dt.strftime("%Y-%m")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Days", len(df))
    c2.metric("Range", f"{df['date'].min().date()} → {df['date'].max().date()}")
    c3.metric("Mode Sum", f"Sum {int(df['sum'].mode().iloc[0])}")
    c4.metric("Mode Jodi", df['jodi'].mode().iloc[0])

    # Per-market Digit 0-9 total count
    st.subheader("🔢 Digit 0-9 Total Count (this market)")
    dc_m = [0]*10
    for j in df["jodi"]:
        dc_m[int(j[0])] += 1
        dc_m[int(j[1])] += 1
    grand_m = sum(dc_m) or 1
    hot_m = dc_m.index(max(dc_m)); cold_m = dc_m.index(min(dc_m))
    cols = st.columns(11)
    cols[0].markdown("**Digit**")
    for d in range(10):
        cols[d+1].markdown(f"<div style='text-align:center;background:{DIGIT_COLORS[d]};color:white;padding:6px;border-radius:5px;font-weight:bold'>{d}</div>", unsafe_allow_html=True)
    cols = st.columns(11)
    cols[0].markdown("**Count**")
    for d in range(10):
        cols[d+1].markdown(f"<div style='text-align:center;font-weight:bold;padding:4px'>{dc_m[d]}</div>", unsafe_allow_html=True)
    cols = st.columns(11)
    cols[0].markdown("**%**")
    for d in range(10):
        cols[d+1].markdown(f"<div style='text-align:center;color:#888;padding:2px'>{dc_m[d]/grand_m*100:.1f}%</div>", unsafe_allow_html=True)
    cA, cB, cC = st.columns(3)
    cA.metric("🔥 Hottest", f"Digit {hot_m}", f"{dc_m[hot_m]} plays")
    cB.metric("❄️ Coldest", f"Digit {cold_m}", f"{dc_m[cold_m]} plays")
    cC.metric("📊 Total digit plays", grand_m)

    st.subheader("Monthly Heatmap")
    pivot = df.groupby(["month","sum"]).size().unstack(fill_value=0)
    for sg in [10,11,12,13,14]:
        if sg not in pivot.columns: pivot[sg] = 0
    pivot = pivot[[10,11,12,13,14]]
    fig = px.imshow(pivot.T, color_continuous_scale="YlOrRd",
                   labels=dict(x="Month", y="Sum Group", color="Count"))
    fig.update_layout(height=300)
    st.plotly_chart(fig, use_container_width=True)
    st.subheader("Calendar View")
    months = sorted(df["month"].unique(), reverse=True)
    mo = st.selectbox("Month", months)
    sub = df[df["month"]==mo].copy(); sub["day"] = sub["date"].dt.day
    year, month = int(mo[:4]), int(mo[5:7])
    cal = calendar.monthcalendar(year, month)
    grid = "<table style='border-collapse:collapse;width:100%'>"
    grid += "<tr>" + "".join(f"<th style='padding:8px;border:1px solid #ccc;background:#305496;color:white'>{d}</th>" for d in ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"]) + "</tr>"
    dmap = {row["day"]: row for _, row in sub.iterrows()}
    for week in cal:
        grid += "<tr>"
        for d in week:
            if d == 0:
                grid += "<td style='padding:8px;border:1px solid #ccc'></td>"
            elif d in dmap:
                row = dmap[d]; sg = int(row["sum"]); color = SUM_COLORS[sg]
                grid += f"<td style='padding:8px;border:1px solid #ccc;background:{color};color:white;text-align:center'><b>{d}</b><br>{row['jodi']}<br><small>Sum {sg}</small></td>"
            else:
                grid += f"<td style='padding:8px;border:1px solid #ccc;color:#aaa;text-align:center'>{d}</td>"
        grid += "</tr>"
    grid += "</table>"
    st.markdown(grid, unsafe_allow_html=True)
    st.subheader("Top 10 Streaks")
    df_s = df.sort_values("date").reset_index(drop=True)
    df_s["run"] = (df_s["sum"] != df_s["sum"].shift()).cumsum()
    streaks = df_s.groupby(["run","sum"]).size().reset_index(name="length")
    streaks = streaks.sort_values("length", ascending=False).head(10)
    streaks.columns = ["Run #","Sum Group","Length"]
    st.dataframe(streaks, use_container_width=True, hide_index=True)
    st.subheader("Full History (sortable)")
    st.dataframe(df[["date","jodi","sum"]].rename(columns={"date":"Date","jodi":"Jodi","sum":"Sum"}).sort_values("Date", ascending=False),
                use_container_width=True, hide_index=True)

def page_settings():
    st.title("⚙️ Settings")
    tabs = st.tabs(["🏪 Markets","👤 Users","🔐 Password","💾 Backup","🔄 Auto-Fetch","📋 Audit Log","📱 PWA"])
    with tabs[0]:
        st.subheader("Markets")
        md = list_markets()
        st.dataframe(md, use_container_width=True, hide_index=True)
        with st.form("am"):
            n = st.text_input("Name")
            cl = st.text_input("Closures (0=Sun..6=Sat)", placeholder="0 = closed Sun")
            ok = st.form_submit_button("Add")
        if ok and n: add_market(n, cl); st.success("Added"); st.rerun()
    with tabs[1]:
        st.subheader("Users")
        if st.session_state.get("role") != "admin":
            st.warning("Admin only.")
        else:
            users = pd.read_sql("SELECT id, username, role, created_at FROM users", get_conn())
            st.dataframe(users, use_container_width=True, hide_index=True)
            with st.form("au"):
                c1, c2, c3 = st.columns(3)
                with c1: nu = st.text_input("Username")
                with c2: pw = st.text_input("Password", type="password")
                with c3: rl = st.selectbox("Role", ["admin","viewer"])
                ok = st.form_submit_button("Add User")
            if ok and nu and pw:
                try:
                    conn = get_conn()
                    conn.execute("INSERT INTO users (username, password_hash, role) VALUES (?,?,?)",
                                 (nu, hash_pw(pw), rl))
                    conn.commit()
                    log_audit("add_user", nu)
                    st.success(f"Added {nu}"); st.rerun()
                except sqlite3.IntegrityError:
                    st.error("Username exists")
    with tabs[2]:
        st.subheader("Change my password")
        with st.form("cp"):
            old = st.text_input("Current", type="password")
            n1 = st.text_input("New", type="password")
            n2 = st.text_input("Confirm", type="password")
            ok = st.form_submit_button("Change")
        if ok:
            row = get_conn().execute("SELECT password_hash FROM users WHERE username=?",
                                     (st.session_state["user"],)).fetchone()
            if not row or row["password_hash"] != hash_pw(old): st.error("Wrong current")
            elif n1 != n2: st.error("Mismatch")
            elif len(n1) < 4: st.error("Too short")
            else:
                conn = get_conn()
                conn.execute("UPDATE users SET password_hash=? WHERE username=?",
                            (hash_pw(n1), st.session_state["user"]))
                conn.commit(); log_audit("change_password")
                st.success("Changed.")
    with tabs[3]:
        st.subheader("Backup / Restore")
        c1, c2 = st.columns(2)
        with c1:
            with open(DB_PATH, "rb") as f: data = f.read()
            st.download_button("⬇️ Download matka.db", data=data,
                              file_name=f"matka_backup_{dt.date.today()}.db",
                              mime="application/octet-stream", use_container_width=True)
            backups = sorted([f for f in os.listdir(BACKUP_DIR) if f.endswith(".db")], reverse=True)
            if backups:
                st.markdown("**Auto-backups (last 10):**")
                st.text("\n".join(backups[:10]))
        with c2:
            st.warning("⚠️ Restore replaces all data.")
            up = st.file_uploader("Restore .db", type=["db"])
            if up and st.button("Restore", type="primary"):
                shutil.copy(DB_PATH, f"{BACKUP_DIR}/before_restore_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.db")
                with open(DB_PATH, "wb") as f: f.write(up.read())
                log_audit("restore_db", up.name)
                st.success("Restored. Reload page."); st.balloons()
    with tabs[4]:
        st.subheader("🔄 Auto-Fetch from source website")
        st.caption("Pulls today's jodis from a results website and saves into the database. "
                   "Useful for daily auto-updates. ⚠️ Check source's Terms of Service before scheduling.")
        try:
            from scraper import run_fetch, DEFAULT_SOURCE_URL
            scraper_available = True
        except Exception as e:
            st.error(f"scraper.py missing or has error: {e}")
            scraper_available = False
            DEFAULT_SOURCE_URL = ""
        if scraper_available:
            # Config persisted in a tiny key/val table
            conn = get_conn()
            conn.execute("""CREATE TABLE IF NOT EXISTS config (
                key TEXT PRIMARY KEY, value TEXT)""")
            conn.commit()
            saved_url = conn.execute("SELECT value FROM config WHERE key='source_url'").fetchone()
            cur_url = saved_url["value"] if saved_url else DEFAULT_SOURCE_URL
            new_url = st.text_input("Source website URL", value=cur_url,
                                   help="Currently targets dpboss-style HTML tables.")
            if new_url != cur_url:
                conn.execute("INSERT OR REPLACE INTO config (key, value) VALUES ('source_url', ?)", (new_url,))
                conn.commit()
                st.success("URL updated.")
            c1, c2, c3 = st.columns([1,1,2])
            with c1:
                fetch_date = st.date_input("Fetch for date", value=dt.date.today(), key="fdate")
            with c2:
                if st.button("🔄 Fetch Now", type="primary"):
                    with st.spinner(f"Fetching from {new_url}..."):
                        res = run_fetch(fetch_date.isoformat(), new_url)
                    if res["ok"]:
                        st.success(f"✓ {res['message']}")
                        if res.get("results"):
                            st.write("**Markets found:**")
                            st.json(res["results"])
                    else:
                        st.error(f"✗ {res['status']}: {res['message']}")
            with c3:
                st.markdown("**Schedule daily auto-fetch (Windows):**\n"
                           "Open Task Scheduler → Create Basic Task → "
                           "Trigger: Daily at preferred time → "
                           "Action: Start a program → "
                           f"Program: `{os.path.abspath('fetch_now.bat')}`")
            # Fetch log
            try:
                log = pd.read_sql(
                    "SELECT ts, date, status, markets_found, markets_saved, message, url FROM fetch_log ORDER BY id DESC LIMIT 50",
                    conn)
                if not log.empty:
                    st.markdown("**Recent fetch history:**")
                    st.dataframe(log, use_container_width=True, hide_index=True, height=300)
                else:
                    st.info("No fetches yet. Click 'Fetch Now' to test.")
            except Exception:
                st.info("No fetch history yet.")
    with tabs[5]:
        st.subheader("Audit Log (filter & search)")
        conn = get_conn()
        c1, c2, c3 = st.columns(3)
        with c1:
            users = pd.read_sql("SELECT DISTINCT user FROM audit_log", conn)["user"].tolist()
            fu = st.selectbox("User", ["(any)"] + users)
        with c2:
            acts = pd.read_sql("SELECT DISTINCT action FROM audit_log", conn)["action"].tolist()
            fa = st.selectbox("Action", ["(any)"] + acts)
        with c3:
            search = st.text_input("Search details")
        q = "SELECT ts, user, action, details FROM audit_log WHERE 1=1"
        params = []
        if fu != "(any)": q += " AND user=?"; params.append(fu)
        if fa != "(any)": q += " AND action=?"; params.append(fa)
        if search: q += " AND details LIKE ?"; params.append(f"%{search}%")
        q += " ORDER BY id DESC LIMIT 500"
        log = pd.read_sql(q, conn, params=params)
        st.dataframe(log, use_container_width=True, hide_index=True, height=400)
    with tabs[6]:
        st.subheader("Install as App (PWA)")
        st.markdown("""
        **iPhone (Safari):** Open the site → Share button → "Add to Home Screen"

        **Android (Chrome):** Open the site → menu → "Install app" or "Add to Home Screen"

        **Windows (Edge/Chrome):** click the install icon in the address bar (small ⊕ icon).

        After install, the app opens fullscreen with no browser bars and has its own icon.
        """)
        if os.path.exists("static/icon-192.png"):
            st.image("static/icon-192.png", width=128)

# ====================================================================
# MAIN
# ====================================================================
init_db()
auto_backup()
if not login_panel(): st.stop()

with st.sidebar:
    st.title("🎯 Matka Tracker v3")
    st.caption(f"👤 {st.session_state['user']} ({st.session_state.get('role','?')})")
    with st.expander("📱 Install as app", expanded=False):
        st.markdown("""
**Android/Chrome:** menu (⋮) → Install app
**iPhone/Safari:** Share → Add to Home Screen
**Windows/Edge:** address bar ⊕ icon → Install

See **PACKAGE_AS_APP.md** for APK/MSIX builds.
""")
    page = st.radio("Navigate", [
        "📊 Dashboard","📝 Data Entry","🔮 Predictions","🧪 Backtest",
        "🧩 Patterns","📑 Reports","📈 Market History","⚙️ Settings"
    ])
    st.divider()
    if st.button("🚪 Logout", use_container_width=True):
        log_audit("logout"); st.session_state.clear(); st.rerun()

if page.endswith("Dashboard"): page_dashboard()
elif page.endswith("Data Entry"): page_data_entry()
elif page.endswith("Predictions"): page_predictions()
elif page.endswith("Backtest"): page_backtest()
elif page.endswith("Patterns"): page_patterns()
elif page.endswith("Reports"): page_reports()
elif page.endswith("Market History"): page_market_history()
elif page.endswith("Settings"): page_settings()
